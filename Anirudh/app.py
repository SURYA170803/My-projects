from flask import Flask, render_template, request, redirect, url_for, flash
import pytesseract
from PIL import Image
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# Path to Tesseract
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Initialize or load Excel file
EXCEL_FILE = 'D:\Anirudh\excel_data.xlsx'
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['Scanned Text'])  # Add header
    wb.save(EXCEL_FILE)


@app.route('/')
def index():
    return render_template('login.html')


@app.route('/login', methods=['POST'])
def login():
    username = request.form['username']
    password = request.form['password']
    
    # Dummy authentication (replace with actual validation logic)
    if username == 'admin' and password == 'password':
        return redirect(url_for('ocr'))
    else:
        flash('Invalid username or password')
        return redirect(url_for('index'))


@app.route('/ocr')
def ocr():
    return render_template('ocr.html')


@app.route('/perform-ocr', methods=['POST'])
def perform_ocr():
    if 'image' not in request.files:
        flash('No file uploaded')
        return redirect(url_for('ocr'))
    
    image = request.files['image']
    if image.filename == '':
        flash('No file selected')
        return redirect(url_for('ocr'))

    image_path = os.path.join('uploads', image.filename)
    os.makedirs('uploads', exist_ok=True)
    image.save(image_path)

    # Perform OCR
    text = pytesseract.image_to_string(Image.open(image_path))

    # Save result to Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([text])
    wb.save(EXCEL_FILE)

    flash('OCR performed and saved to Excel')
    return redirect(url_for('ocr'))


if __name__ == '__main__':
    app.run(debug=True)
